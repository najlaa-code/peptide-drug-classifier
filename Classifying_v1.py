import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook

class PeptideAnalyzer:
    def __init__(self, excel_file_path):
        """Initialize the excel file path"""
        self.file_path = Path(excel_file_path)
        self.df=None
        # Industry standard thresholds
        self.MIC_THRESHOLD=1.0
        self.HC50_THRESHOLD=1.7 #log base 10 to 50
    def load_data(self):
        """Load and validate the excel data, exception handling"""
        try:
            self.df = pd.read_excel(self.file_path)
            # if need to print all the peptides
            #print(f"Loaded peptides... : {len(self.df)} ")

            #if need to verify if the columns are here, add code here. In our case with the clean database, we already know it's there
        except Exception as e:
            raise Exception(f"Failed to load data/access the csv file. {str(e)}")
    def classify_peptides(self):
        """Based on the 4 classes previously defined, time to classify them to get an overview."""
        if self.df is None:
            raise ValueError("No data loaded.")
        conditions = [
            # Class 0: Selective (Active + Non-toxic)
            (self.df['log10mic'] <= self.MIC_THRESHOLD) & (self.df['log10hc50'] > self.HC50_THRESHOLD),
            # Class 1: Pure Hemolytic (Inactive + Toxic)
            (self.df['log10mic'] > self.MIC_THRESHOLD) & (self.df['log10hc50'] <= self.HC50_THRESHOLD),
            # Class 2: Both Active and Toxic
            (self.df['log10mic'] <= self.MIC_THRESHOLD) & (self.df['log10hc50'] <= self.HC50_THRESHOLD),
            # Class 3: Inactive (Inactive + Non-toxic)
            (self.df['log10mic'] > self.MIC_THRESHOLD) & (self.df['log10hc50'] > self.HC50_THRESHOLD)
        ]
        labels = [
            'Class 0: Selective',
            'Class 1: Pure Hemolytic',
            'Class 2: Both Active/Toxic',
            'Class 3: Inactive'
        ]
        self.df['drug_class'] = np.select(conditions, labels, default='Unclassified') # Add a column in the excel file mentionning which was one is classified.
        # if need to add active/inactive/toxic/nontoxic do it here

        # --- RUN 1: FULL DATASET (n_hemo_measurements < 0)
        print("Full dataset classification:")
        # to know how many peptides in each class
        class_counts = self.df['drug_class'].value_counts()
        total = len(self.df)
        for class_name, count in class_counts.items():
            percentage = (count/total)*100
            print(f"{class_name:25} {count:5d}({percentage:5f}%)")
        print(f"{'Total peptides':25}{total:5d}")

        # --- RUN 2: FILTERED DATASET (n_hemo_measurements > 1)
        filtered = self.df['n_hemo_measurements'] > 1
        self.df['drug_class_filtered'] = np.where(
            filtered,
            np.select(conditions, labels, default='Unclassified'),
            'N/A'
        )
        print("Filtered dataset classification: (only includes n_hemo_measurement > 1 only)")
        filtered_df = self.df[filtered]
        filtered_total = len(filtered_df)
        for class_name, count in filtered_df['drug_class_filtered'].value_counts().items():
            percentage = (count/filtered_total)*100
            print(f"{class_name:25} {count:5d}({percentage:5f}%)")
        print(f"{'Total peptides':25}{filtered_total:5d}")
        print(f"{'N/A (excluded)':25} {total - filtered_total:5d}")


        return class_counts
    def export_results(self):
            """Show the results of the analysis in a column in the excel file"""
            wb = load_workbook(self.file_path)
            ws = wb.active
            # Add headers for the columns by finding the empty column, then adding the header names. Note: row 1 is the header row
            next_column = ws.max_column + 1
            ws.cell(row=1, column=next_column, value='drug_class')
            ws.cell(row=1, column=next_column + 1, value='drug_class_filtered')
            # write corresponding class/value starting from row 2
            for i, (_, row) in enumerate(self.df.iterrows(), start=2):
                ws.cell(row=i, column=next_column, value=row['drug_class'])
                ws.cell(row=i, column=next_column + 1, value=row['drug_class_filtered'])
             # save it
            wb.save(self.file_path)
    def generate_results(self):
            """This method runs all the required analysis to classify the peptides"""
            self.load_data()
            self.classify_peptides()
            self.export_results()
            print("End of analysis")

def main():
        """Main function to run the analysis"""
        import sys
        # Check if the file dbaasp_grampa_hemolytik_hc50_mic.csv exists
        excel_file = sys.argv[1]
        if not Path(excel_file).exists():
            print(f"Error: File '{excel_file}' not found.")
            sys.exit(1)
        # Start the analysis
        analyzer = PeptideAnalyzer(excel_file)
        analyzer.generate_results()

if __name__ == "__main__":
    main()







